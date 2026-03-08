from flask import Blueprint, request, jsonify, send_file, Response
from extensions import db
from flask_jwt_extended import jwt_required, get_jwt_identity
from datetime import datetime
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable, KeepTogether
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm, mm
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
from reportlab.platypus import Image as RLImage
from PIL import Image as PILImage
from io import BytesIO
import os, re

# ── openpyxl for Excel export ──────────────────────────────────────────────────
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
from openpyxl.utils import get_column_letter

quotation_bp = Blueprint('quotation', __name__)

# ── Model ──────────────────────────────────────────────────────────────────────
class Quotation(db.Model):
    __tablename__ = "quotations"
    id               = db.Column(db.Integer, primary_key=True)
    quotation_number = db.Column(db.String(60), unique=True)
    base_number      = db.Column(db.String(30))
    revision         = db.Column(db.Integer, default=0)
    customer_name    = db.Column(db.String(150))
    customer_company = db.Column(db.String(200))
    customer_email   = db.Column(db.String(120))
    customer_phone   = db.Column(db.String(30))
    customer_address = db.Column(db.Text)
    project_name     = db.Column(db.String(200))
    category         = db.Column(db.String(100))
    status           = db.Column(db.String(30), default="draft")
    valid_until      = db.Column(db.Date)
    currency         = db.Column(db.String(10), default="IDR")
    total_amount     = db.Column(db.Float, default=0)
    notes            = db.Column(db.Text)
    terms            = db.Column(db.Text)
    items            = db.Column(db.JSON)
    sales_person     = db.Column(db.String(100))
    ref_no           = db.Column(db.String(100))
    shipment_terms   = db.Column(db.String(200))
    delivery         = db.Column(db.String(200))
    payment_terms    = db.Column(db.String(200))
    vat_pct          = db.Column(db.Float, default=11.0)
    vat_include      = db.Column(db.Boolean, default=False)
    created_by       = db.Column(db.Integer, db.ForeignKey("users.id"))
    created_at       = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at       = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

# ── Constants ──────────────────────────────────────────────────────────────────
FLOTECH_INFO = {
    "name":    "PT. FLOTECH CONTROLS INDONESIA",
    "address": "Rukan Artha Gading Niaga, Blok F/7",
    "city":    "Jl. Boulevard Artha Gading, Jakarta 14240",
    "telp":    "Telp: (021) 4585 0778 / 4586 0625  |  Fax: (021) 4585 0779",
    "email":   "Email: salesjkt@flotech.co.id  |  Website: www.flotech.co.id",
}

# ── Helpers ────────────────────────────────────────────────────────────────────
def format_currency(amount, currency="IDR"):
    try:
        n = float(amount or 0)
        if currency == "IDR":
            return f"Rp {n:,.0f}".replace(",", ".")
        return f"{currency} {n:,.2f}"
    except:
        return "-"

def calc_item(item):
    price = float(item.get("unit_price") or 0)
    qty   = float(item.get("qty") or 0)
    disc  = float(item.get("discount") or 0)
    gross = price * qty
    disc_amt = gross * disc / 100
    net   = gross - disc_amt
    return gross, disc_amt, net

def _find_logo():
    base = os.path.dirname(os.path.abspath(__file__))
    candidates = [
        os.path.join(base, "..", "frontend", "public", "logo.png"),
        os.path.join(base, "static", "logo.png"),
        os.path.join(base, "static", "flotech_logo.png"),
        os.path.join(base, "..", "..", "frontend", "public", "logo.png"),
    ]
    for p in candidates:
        n = os.path.normpath(p)
        if os.path.isfile(n):
            return n
    return None

# ── Auto-number generator ──────────────────────────────────────────────────────
def generate_quotation_number():
    now = datetime.utcnow()
    yy, mm_str = now.strftime("%y"), now.strftime("%m")
    prefix = f"SQ{yy}{mm_str}"
    year_prefix = f"SQ{yy}"
    existing = Quotation.query.filter(
        Quotation.base_number.like(f"{year_prefix}%")
    ).order_by(Quotation.id.desc()).all()
    max_seq = 0
    for q in existing:
        if q.base_number:
            try:
                seq = int(q.base_number[6:])
                if seq > max_seq:
                    max_seq = seq
            except:
                pass
    return f"{prefix}{max_seq + 1:03d}"

# ═══════════════════════════════════════════════════════════════════════════════
# ROUTES
# ═══════════════════════════════════════════════════════════════════════════════
@quotation_bp.route('/next-number', methods=['GET'])
@jwt_required()
def get_next_number():
    return jsonify({"number": generate_quotation_number()}), 200

@quotation_bp.route('/list', methods=['GET'])
@jwt_required()
def list_quotations():
    qs = Quotation.query.order_by(Quotation.created_at.desc()).all()
    result = []
    for q in qs:
        result.append({
            "id": q.id, "quotation_number": q.quotation_number,
            "base_number": q.base_number, "revision": q.revision or 0,
            "customer_name": q.customer_name, "customer_company": q.customer_company,
            "project_name": q.project_name, "category": q.category,
            "status": q.status, "total_amount": q.total_amount, "currency": q.currency,
            "sales_person": q.sales_person,
            "created_at": q.created_at.isoformat() if q.created_at else None,
            "updated_at": q.updated_at.isoformat() if q.updated_at else None,
            "valid_until": q.valid_until.isoformat() if q.valid_until else None,
        })
    return jsonify(result), 200

@quotation_bp.route('/analytics', methods=['GET'])
@jwt_required()
def get_analytics():
    qs = Quotation.query.all()
    monthly, yearly = {}, {}
    for q in qs:
        dt = q.created_at or datetime.utcnow()
        mk, yk = dt.strftime("%Y-%m"), dt.strftime("%Y")
        for store, key in [(monthly, mk), (yearly, yk)]:
            if key not in store:
                store[key] = {"period": key, "draft": 0, "sent": 0, "followup": 0, "won": 0, "lost": 0, "cancel": 0,
                              "draft_val": 0, "sent_val": 0, "followup_val": 0, "won_val": 0, "lost_val": 0, "cancel_val": 0, "total": 0}
            st = q.status or "draft"
            val = q.total_amount or 0
            store[key][st] = store[key].get(st, 0) + 1
            store[key][f"{st}_val"] = store[key].get(f"{st}_val", 0) + val
            store[key]["total"] += 1
    return jsonify({
        "monthly": sorted(monthly.values(), key=lambda x: x["period"]),
        "yearly":  sorted(yearly.values(),  key=lambda x: x["period"]),
    }), 200

@quotation_bp.route('/create', methods=['POST'])
@jwt_required()
def create_quotation():
    user_id = int(get_jwt_identity())
    data = request.get_json()
    valid_until = None
    if data.get("valid_until"):
        try: valid_until = datetime.strptime(data["valid_until"], "%Y-%m-%d").date()
        except: pass
    base_num = data.get("base_number") or generate_quotation_number()
    q = Quotation(
        quotation_number=base_num, base_number=base_num, revision=0,
        customer_name=data.get("customer_name"), customer_company=data.get("customer_company"),
        customer_email=data.get("customer_email"), customer_phone=data.get("customer_phone"),
        customer_address=data.get("customer_address"), project_name=data.get("project_name"),
        category=data.get("category"), status=data.get("status", "draft"),
        valid_until=valid_until, currency=data.get("currency", "IDR"),
        total_amount=data.get("total_amount", 0), notes=data.get("notes"), terms=data.get("terms"),
        items=data.get("items", []), sales_person=data.get("sales_person"), ref_no=data.get("ref_no"),
        shipment_terms=data.get("shipment_terms"), delivery=data.get("delivery"),
        payment_terms=data.get("payment_terms"),
        vat_pct=float(data.get("vat_pct") or 11), vat_include=bool(data.get("vat_include", False)),
        created_by=user_id,
    )
    db.session.add(q); db.session.commit()
    return jsonify({"message": "Created", "id": q.id, "quotation_number": q.quotation_number}), 201

@quotation_bp.route('/detail/<int:qid>', methods=['GET'])
@jwt_required()
def get_quotation(qid):
    q = Quotation.query.get(qid)
    if not q: return jsonify({"error": "Not found"}), 404
    return jsonify({
        "id": q.id, "quotation_number": q.quotation_number,
        "base_number": q.base_number, "revision": q.revision or 0,
        "customer_name": q.customer_name, "customer_company": q.customer_company,
        "customer_email": q.customer_email, "customer_phone": q.customer_phone,
        "customer_address": q.customer_address, "project_name": q.project_name,
        "category": q.category, "status": q.status,
        "valid_until": q.valid_until.isoformat() if q.valid_until else None,
        "currency": q.currency, "total_amount": q.total_amount,
        "notes": q.notes, "terms": q.terms, "items": q.items or [],
        "sales_person": q.sales_person, "ref_no": q.ref_no,
        "shipment_terms": q.shipment_terms, "delivery": q.delivery,
        "payment_terms": q.payment_terms,
        "vat_pct": q.vat_pct or 11, "vat_include": q.vat_include or False,
        "created_at": q.created_at.isoformat() if q.created_at else None,
        "updated_at": q.updated_at.isoformat() if q.updated_at else None,
    }), 200

@quotation_bp.route('/update/<int:qid>', methods=['PUT'])
@jwt_required()
def update_quotation(qid):
    q = Quotation.query.get(qid)
    if not q: return jsonify({"error": "Not found"}), 404
    data = request.get_json()
    REVISION_FIELDS = {"items", "total_amount", "notes", "terms", "payment_terms",
                       "shipment_terms", "delivery", "currency"}
    bump = data.get("bump_revision", False)
    if not bump:
        for rf in REVISION_FIELDS:
            if rf in data and str(getattr(q, rf, None)) != str(data[rf]):
                bump = True; break
    for f in ["customer_name","customer_company","customer_email","customer_phone","customer_address",
              "project_name","category","currency","notes","terms","items","total_amount","status",
              "sales_person","ref_no","shipment_terms","delivery","payment_terms","vat_pct","vat_include"]:
        if f in data: setattr(q, f, data[f])
    if data.get("valid_until"):
        try: q.valid_until = datetime.strptime(data["valid_until"], "%Y-%m-%d").date()
        except: pass
    if bump:
        new_rev = (q.revision or 0) + 1
        q.revision = new_rev
        base = q.base_number or q.quotation_number
        q.base_number = base
        q.quotation_number = f"{base}-Rev{new_rev}"
    q.updated_at = datetime.utcnow()
    db.session.commit()
    return jsonify({"message": "Updated", "quotation_number": q.quotation_number, "revision": q.revision}), 200

@quotation_bp.route('/status/<int:qid>', methods=['PUT'])
@jwt_required()
def update_status(qid):
    q = Quotation.query.get(qid)
    if not q: return jsonify({"error": "Not found"}), 404
    q.status = request.get_json().get("status", q.status)
    q.updated_at = datetime.utcnow()
    db.session.commit()
    return jsonify({"message": "Updated"}), 200

@quotation_bp.route('/delete/<int:qid>', methods=['DELETE'])
@jwt_required()
def delete_quotation(qid):
    q = Quotation.query.get(qid)
    if not q: return jsonify({"error": "Not found"}), 404
    db.session.delete(q); db.session.commit()
    return jsonify({"message": "Deleted"}), 200

# ═══════════════════════════════════════════════════════════════════════════════
# EXPORT — Excel list of filtered quotations
# ═══════════════════════════════════════════════════════════════════════════════
@quotation_bp.route('/export/excel', methods=['POST'])
@jwt_required()
def export_excel():
    """Receive list of quotation IDs (or all) and return XLSX."""
    data = request.get_json() or {}
    ids  = data.get("ids")  # list of IDs to export; None = all

    if ids:
        qs = Quotation.query.filter(Quotation.id.in_(ids)).order_by(Quotation.created_at.desc()).all()
    else:
        qs = Quotation.query.order_by(Quotation.created_at.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Quotations"

    # ── colours ────────────────────────────────────────────────────────────────
    navy   = "0B3D91"
    navy2  = "1E5CC6"
    accent = "E8EEF9"
    green  = "D1FAE5"
    amber  = "FEF3C7"
    red    = "FEE2E2"
    gray   = "F9FAFB"
    white  = "FFFFFF"

    status_fill = {
        "draft":    "F3F4F6", "sent":     "DBEAFE", "followup": "FEF3C7",
        "won":      "D1FAE5", "lost":     "FEE2E2", "cancel":   "F3F4F6",
    }
    status_fc = {
        "draft": "374151", "sent": "1D4ED8", "followup": "92400E",
        "won": "065F46", "lost": "991B1B", "cancel": "6B7280",
    }

    def hf(hex_):
        return PatternFill("solid", fgColor=hex_)

    def bf(bold=False, size=9, color="1F2937", name="Arial"):
        return Font(name=name, size=size, bold=bold, color=color)

    thin = Side(style="thin", color="D1D5DB")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    border_btm = Border(bottom=Side(style="medium", color="0B3D91"))

    # ── Title block ────────────────────────────────────────────────────────────
    ws.merge_cells("A1:K1")
    ws["A1"] = "PT. FLOTECH CONTROLS INDONESIA"
    ws["A1"].font = Font(name="Arial", size=14, bold=True, color=white)
    ws["A1"].fill = hf(navy)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:K2")
    ws["A2"] = "Sales Quotation Report"
    ws["A2"].font = Font(name="Arial", size=10, bold=False, color="BFDBFE")
    ws["A2"].fill = hf(navy2)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    ws.merge_cells("A3:K3")
    ws["A3"] = f"Generated: {datetime.utcnow().strftime('%d %B %Y %H:%M')} UTC  |  Total: {len(qs)} quotations"
    ws["A3"].font = Font(name="Arial", size=8, italic=True, color="6B7280")
    ws["A3"].fill = hf("F8FAFC")
    ws["A3"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 14

    ws.row_dimensions[4].height = 6  # spacer

    # ── Headers ────────────────────────────────────────────────────────────────
    headers = [
        "No.", "No. Quotation", "Customer / Perusahaan", "Nama PIC",
        "Project / Subject", "Sales Person", "Currency",
        "Total Nilai", "Status", "Berlaku s/d", "Dibuat"
    ]
    header_widths = [5, 20, 30, 22, 35, 20, 10, 18, 12, 14, 18]

    for col_idx, (h, w) in enumerate(zip(headers, header_widths), start=1):
        cell = ws.cell(row=5, column=col_idx, value=h)
        cell.font = Font(name="Arial", size=9, bold=True, color=white)
        cell.fill = hf(navy)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_all
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.row_dimensions[5].height = 24

    # ── Data rows ──────────────────────────────────────────────────────────────
    for i, q in enumerate(qs):
        row = 6 + i
        bg = gray if i % 2 == 0 else white
        st = (q.status or "draft").lower()

        vals = [
            i + 1,
            q.quotation_number,
            q.customer_company or "",
            q.customer_name or "",
            q.project_name or "",
            q.sales_person or "",
            q.currency or "IDR",
            q.total_amount or 0,
            (q.status or "draft").title(),
            q.valid_until.strftime("%d-%b-%Y") if q.valid_until else "",
            q.created_at.strftime("%d-%b-%Y") if q.created_at else "",
        ]

        for col_idx, val in enumerate(vals, start=1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = bf(size=9)
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            cell.border = border_all

            # Column-specific formatting
            if col_idx == 1:  # No.
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = hf(bg)
            elif col_idx == 2:  # Quotation number
                cell.font = Font(name="Arial", size=9, bold=True, color=navy)
                cell.fill = hf(bg)
            elif col_idx == 8:  # Total nilai
                cell.number_format = '#,##0' if (q.currency or "IDR") == "IDR" else '#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.fill = hf(bg)
            elif col_idx == 9:  # Status
                sfill = status_fill.get(st, "F3F4F6")
                sfc   = status_fc.get(st, "374151")
                cell.fill = hf(sfill)
                cell.font = Font(name="Arial", size=9, bold=True, color=sfc)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.fill = hf(bg)

        ws.row_dimensions[row].height = 18

    # ── Summary section ────────────────────────────────────────────────────────
    summary_row = 6 + len(qs) + 2

    ws.merge_cells(f"A{summary_row}:G{summary_row}")
    ws[f"A{summary_row}"] = "RINGKASAN"
    ws[f"A{summary_row}"].font = Font(name="Arial", size=9, bold=True, color=white)
    ws[f"A{summary_row}"].fill = hf(navy)
    ws[f"A{summary_row}"].alignment = Alignment(horizontal="center")

    status_labels = [("Won","won","065F46",green), ("Sent","sent","1D4ED8","DBEAFE"),
                     ("Follow Up","followup","92400E",amber), ("Draft","draft","374151",gray),
                     ("Lost","lost","991B1B",red)]
    for j, (label, key, fc, bg_hex) in enumerate(status_labels):
        r = summary_row + 1 + j
        cnt  = sum(1 for q in qs if (q.status or "draft") == key)
        val  = sum((q.total_amount or 0) for q in qs if (q.status or "draft") == key)
        ws.merge_cells(f"A{r}:C{r}")
        ws[f"A{r}"] = label
        ws[f"A{r}"].font  = Font(name="Arial", size=9, bold=True, color=fc)
        ws[f"A{r}"].fill  = hf(bg_hex)
        ws[f"A{r}"].alignment = Alignment(horizontal="left")
        ws[f"D{r}"] = cnt
        ws[f"D{r}"].font = Font(name="Arial", size=9, bold=True, color=fc)
        ws[f"D{r}"].fill = hf(bg_hex)
        ws[f"D{r}"].alignment = Alignment(horizontal="center")
        ws[f"E{r}"] = val
        ws[f"E{r}"].number_format = '#,##0'
        ws[f"E{r}"].font = Font(name="Arial", size=9, color=fc)
        ws[f"E{r}"].fill = hf(bg_hex)
        ws[f"E{r}"].alignment = Alignment(horizontal="right")

    # ── Freeze panes ───────────────────────────────────────────────────────────
    ws.freeze_panes = "A6"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"Quotations_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ═══════════════════════════════════════════════════════════════════════════════
# EXPORT — PDF list of filtered quotations
# ═══════════════════════════════════════════════════════════════════════════════
@quotation_bp.route('/export/pdf', methods=['POST'])
@jwt_required()
def export_pdf_list():
    data = request.get_json() or {}
    ids  = data.get("ids")
    if ids:
        qs = Quotation.query.filter(Quotation.id.in_(ids)).order_by(Quotation.created_at.desc()).all()
    else:
        qs = Quotation.query.order_by(Quotation.created_at.desc()).all()

    buffer = BytesIO()
    L = R = 1.8 * cm
    usable = A4[0] - L - R  # ~17.4 cm
    doc = SimpleDocTemplate(buffer, pagesize=A4,
        topMargin=2*cm, bottomMargin=3.5*cm, leftMargin=L, rightMargin=R,
        title="Quotation List Report")

    primary = colors.HexColor("#0B3D91")
    navy2   = colors.HexColor("#1E5CC6")
    accent  = colors.HexColor("#E8EEF9")
    gray    = colors.HexColor("#F9FAFB")
    dark    = colors.HexColor("#1F2937")
    text_c  = colors.HexColor("#374151")
    gray_c  = colors.HexColor("#6B7280")
    white   = colors.white
    border  = colors.HexColor("#D1D5DB")
    status_fill = {
        "draft": colors.HexColor("#F3F4F6"), "sent": colors.HexColor("#DBEAFE"),
        "followup": colors.HexColor("#FEF3C7"), "won": colors.HexColor("#D1FAE5"),
        "lost": colors.HexColor("#FEE2E2"), "cancel": colors.HexColor("#F3F4F6"),
    }
    status_fc = {
        "draft": colors.HexColor("#374151"), "sent": colors.HexColor("#1D4ED8"),
        "followup": colors.HexColor("#92400E"), "won": colors.HexColor("#065F46"),
        "lost": colors.HexColor("#991B1B"), "cancel": colors.HexColor("#6B7280"),
    }

    def ps(name, **kw):
        d = dict(fontName="Helvetica", fontSize=9, leading=12, textColor=text_c)
        d.update(kw); return ParagraphStyle(name, **d)

    elements = []

    # Logo + title
    logo_path = _find_logo()
    logo_cell = ""
    if logo_path:
        try:
            pil = PILImage.open(logo_path)
            w, h = pil.size
            lw = 4.5 * cm; lh = lw * h / w
            logo_cell = RLImage(logo_path, width=lw, height=lh)
        except:
            logo_cell = Paragraph("<b>FLOTECH</b>", ps("lg", fontSize=18, fontName="Helvetica-Bold", textColor=primary))
    else:
        logo_cell = Paragraph("<b>FLOTECH</b>", ps("lg2", fontSize=18, fontName="Helvetica-Bold", textColor=primary))

    title_cell = [
        Paragraph("SALES QUOTATION LIST", ps("t", fontSize=14, fontName="Helvetica-Bold", textColor=primary, leading=18)),
        Paragraph(f"PT. FLOTECH CONTROLS INDONESIA", ps("s", fontSize=9, textColor=gray_c)),
        Paragraph(f"Generated: {datetime.utcnow().strftime('%d %B %Y')}  |  {len(qs)} quotations", ps("s2", fontSize=8, textColor=gray_c)),
    ]
    top = Table([[logo_cell, title_cell]], colWidths=[6*cm, usable-6*cm])
    top.setStyle(TableStyle([
        ('VALIGN', (0,0),(-1,-1), 'MIDDLE'),
        ('ALIGN', (1,0),(1,0), 'RIGHT'),
    ]))
    elements.append(top)
    elements.append(HRFlowable(width=usable, thickness=1, color=primary, spaceAfter=0.3*cm))

    # Table
    th = ps("th", fontSize=8, fontName="Helvetica-Bold", textColor=white, alignment=TA_CENTER)
    td = ps("td", fontSize=8, leading=11)
    td_r = ps("tdr", fontSize=8, leading=11, alignment=TA_RIGHT)
    td_c = ps("tdc", fontSize=8, leading=11, alignment=TA_CENTER)
    td_b = ps("tdb", fontSize=8, fontName="Helvetica-Bold", leading=11, textColor=primary)

    # Cols: No | No.Qt | Company | Project | Sales | Nilai | Status | Tgl
    col_w = [0.6*cm, 3.2*cm, 4.2*cm, 4.0*cm, 2.4*cm, 2.5*cm, 1.7*cm]  # = 18.6cm — slightly over, scale down
    # Recalculate to fit usable exactly
    total_cw = sum(col_w)
    col_w = [w * usable / total_cw for w in col_w]

    tbl_data = [[
        Paragraph("No", th), Paragraph("No. Quotation", th),
        Paragraph("Customer", th), Paragraph("Project", th),
        Paragraph("Sales", th), Paragraph("Nilai", th), Paragraph("Status", th),
    ]]

    row_styles = [
        ('BACKGROUND', (0,0), (-1,0), primary),
        ('TEXTCOLOR',  (0,0), (-1,0), white),
        ('ALIGN',      (0,0), (-1,0), 'CENTER'),
        ('VALIGN',     (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ('LEFTPADDING', (0,0), (-1,-1), 4),
        ('RIGHTPADDING', (0,0), (-1,-1), 4),
        ('INNERGRID',  (0,0), (-1,-1), 0.3, border),
        ('BOX',        (0,0), (-1,-1), 0.5, colors.HexColor("#CBD5E1")),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [white, gray]),
    ]

    for i, q in enumerate(qs):
        st = (q.status or "draft").lower()
        st_label = {"draft":"Draft","sent":"Sent","followup":"Follow Up","won":"Won","lost":"Lost","cancel":"Cancelled"}.get(st, st.title())
        tbl_data.append([
            Paragraph(str(i+1), td_c),
            Paragraph(q.quotation_number or "-", td_b),
            Paragraph(f"<b>{q.customer_company or '-'}</b><br/><font size='7' color='#6B7280'>{q.customer_name or ''}</font>", td),
            Paragraph(q.project_name or "-", td),
            Paragraph(q.sales_person or "-", td),
            Paragraph(format_currency(q.total_amount, q.currency or "IDR"), td_r),
            Paragraph(f"<b>{st_label}</b>",
                      ps(f"st_{i}", fontSize=7, fontName="Helvetica-Bold", textColor=status_fc.get(st, text_c), alignment=TA_CENTER)),
        ])
        row_styles.append(('BACKGROUND', (6, i+1), (6, i+1), status_fill.get(st, gray)))

    tbl = Table(tbl_data, colWidths=col_w, repeatRows=1)
    tbl.setStyle(TableStyle(row_styles))
    elements.append(tbl)
    elements.append(Spacer(1, 0.5*cm))

    # Summary
    total_all = sum((q.total_amount or 0) for q in qs)
    won_all   = sum((q.total_amount or 0) for q in qs if q.status == "won")
    won_cnt   = sum(1 for q in qs if q.status == "won")

    summ_data = [
        [Paragraph("RINGKASAN", ps("sh", fontName="Helvetica-Bold", fontSize=9, textColor=white)), "", "", ""],
        [Paragraph("Total Quotation", ps("sl")), Paragraph(str(len(qs)), ps("sv", fontName="Helvetica-Bold")), "", ""],
        [Paragraph("Total Nilai",     ps("sl")), Paragraph(format_currency(total_all), ps("sv", fontName="Helvetica-Bold")), "", ""],
        [Paragraph("Won",             ps("sl")), Paragraph(f"{won_cnt} qt  |  {format_currency(won_all)}", ps("sv", fontName="Helvetica-Bold", textColor=colors.HexColor("#065F46"))), "", ""],
    ]
    summ_t = Table(summ_data, colWidths=[4*cm, 7*cm, usable-11*cm])
    summ_t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), primary), ('SPAN', (0,0), (-1,0)),
        ('ALIGN', (0,0), (-1,0), 'CENTER'), ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BACKGROUND', (0,1), (-1,-1), accent),
        ('BOX', (0,0), (-1,-1), 0.5, border),
        ('INNERGRID', (0,0), (-1,-1), 0.3, border),
        ('TOPPADDING', (0,0), (-1,-1), 4), ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ('LEFTPADDING', (0,0), (-1,-1), 6), ('RIGHTPADDING', (0,0), (-1,-1), 6),
    ]))
    elements.append(summ_t)

    def footer_cb(cv, doc_obj):
        cv.saveState()
        pw, ph = A4
        cv.setStrokeColor(primary); cv.setLineWidth(1)
        cv.line(L, 3.0*cm, pw-R, 3.0*cm)
        cv.setFont("Helvetica-Bold", 9); cv.setFillColor(primary)
        cv.drawCentredString(pw/2, 2.55*cm, FLOTECH_INFO["name"])
        cv.setFont("Helvetica", 7.5); cv.setFillColor(colors.HexColor("#6B7280"))
        cv.drawCentredString(pw/2, 2.2*cm,  f"{FLOTECH_INFO['address']}  |  {FLOTECH_INFO['city']}")
        cv.drawCentredString(pw/2, 1.85*cm, FLOTECH_INFO["telp"])
        cv.drawCentredString(pw/2, 1.5*cm,  FLOTECH_INFO["email"])
        cv.setFont("Helvetica", 7); cv.setFillColor(colors.HexColor("#9CA3AF"))
        cv.drawRightString(pw-R, 1.1*cm, f"Page {doc_obj.page}")
        cv.restoreState()

    doc.build(elements, onFirstPage=footer_cb, onLaterPages=footer_cb)
    buffer.seek(0)
    filename = f"QuotationList_{datetime.utcnow().strftime('%Y%m%d')}.pdf"
    return send_file(buffer, as_attachment=True, download_name=filename, mimetype="application/pdf")

# ═══════════════════════════════════════════════════════════════════════════════
# PDF BUILDER — Single quotation  (A4, left/right margin 2cm = usable 17cm)
# ═══════════════════════════════════════════════════════════════════════════════
def build_quotation_pdf(q):
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
    buffer  = BytesIO()
    L = R   = 2.0 * cm          # left & right margin
    usable  = 17.0 * cm         # 210mm - 2×20mm = 170mm

    doc = SimpleDocTemplate(buffer, pagesize=A4,
        topMargin    = 2.2 * cm,
        bottomMargin = 4.2 * cm,
        leftMargin   = L,
        rightMargin  = R,
        title  = f"Quotation {q.quotation_number}",
        author = "PT. Flotech Controls Indonesia")

    # ── Colour palette ───────────────────────────────────────────────────────
    primary  = colors.HexColor("#0B3D91")
    navy2    = colors.HexColor("#1E5CC6")
    accent   = colors.HexColor("#EEF3FB")
    dark     = colors.HexColor("#111827")
    text_c   = colors.HexColor("#374151")
    gray_c   = colors.HexColor("#6B7280")
    lt_gray  = colors.HexColor("#F9FAFB")
    border_c = colors.HexColor("#D1D5DB")
    white    = colors.white
    disc_bg  = colors.HexColor("#FFF7ED")
    disc_fc  = colors.HexColor("#C2410C")

    # ── ParagraphStyle factory ───────────────────────────────────────────────
    def ps(name, **kw):
        d = dict(fontName="Helvetica", fontSize=9, leading=12, textColor=text_c)
        d.update(kw)
        return ParagraphStyle(name, **d)

    elements = []

    # ────────────────────────────────────────────────────────────────────────
    # 1.  HEADER  — Logo (left)  +  "SALES QUOTATION" (right)
    # ────────────────────────────────────────────────────────────────────────
    logo_path = _find_logo()
    if logo_path:
        try:
            pil    = PILImage.open(logo_path)
            logo_w = 5.0 * cm
            logo_h = logo_w * pil.size[1] / pil.size[0]
            logo_cell = RLImage(logo_path, width=logo_w, height=logo_h)
        except Exception:
            logo_cell = Paragraph("<b>FLOTECH</b>",
                ps("lf", fontName="Helvetica-Bold", fontSize=20, textColor=primary))
    else:
        logo_cell = Paragraph(
            "<b>FLOTECH</b><br/><font size='7'>PROCESS CONTROL &amp; INSTRUMENTATION</font>",
            ps("lf2", fontName="Helvetica-Bold", fontSize=18, textColor=primary, leading=22))

    right_hdr = [
        Paragraph("PT. FLOTECH CONTROLS INDONESIA",
            ps("cn", fontName="Helvetica-Bold", fontSize=10, textColor=primary,
               leading=13, alignment=TA_RIGHT)),
        Spacer(1, 3),
        Paragraph("SALES QUOTATION",
            ps("qt", fontName="Helvetica-Bold", fontSize=18, textColor=primary,
               leading=22, alignment=TA_RIGHT)),
    ]

    # col widths: logo=6cm, gap implicitly 0, right=11cm  → total 17cm
    hdr_t = Table([[logo_cell, right_hdr]], colWidths=[6.0 * cm, 11.0 * cm])
    hdr_t.setStyle(TableStyle([
        ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN',         (1, 0), (1,  0),  'RIGHT'),
        ('LEFTPADDING',   (0, 0), (-1, -1), 0),
        ('RIGHTPADDING',  (0, 0), (-1, -1), 0),
        ('TOPPADDING',    (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
    ]))
    elements.append(hdr_t)
    elements.append(HRFlowable(width=usable, thickness=1, color=primary, spaceAfter=4))

    # ────────────────────────────────────────────────────────────────────────
    # 2.  META-INFO BAND  — 6 cells spanning full usable width
    # ────────────────────────────────────────────────────────────────────────
    q_date  = q.created_at.strftime("%d %B %Y") if q.created_at else datetime.utcnow().strftime("%d %B %Y")
    rev_str = f"Rev. {q.revision}" if (q.revision or 0) > 0 else "Original"
    cur_str = q.currency or "IDR"

    def meta_cell(lbl, val):
        return [
            Paragraph(lbl, ps(f"ml{lbl}", fontSize=7, fontName="Helvetica-Bold",
                               textColor=gray_c, leading=9)),
            Paragraph(str(val), ps(f"mv{lbl}", fontSize=9, fontName="Helvetica-Bold",
                                    textColor=dark, leading=12)),
        ]

    meta_cols = [
        meta_cell("QUOTATION NO.", q.quotation_number or "-"),
        meta_cell("DATE",          q_date),
        meta_cell("SALES PERSON",  q.sales_person or "-"),
        meta_cell("REVISION",      rev_str),
        meta_cell("REF. NO.",      q.ref_no or "-"),
        meta_cell("CURRENCY",      cur_str),
    ]
    cw6 = usable / 6   # ≈ 2.833 cm each

    meta_t = Table(
        [[c[0] for c in meta_cols],
         [c[1] for c in meta_cols]],
        colWidths=[cw6] * 6,
    )
    meta_t.setStyle(TableStyle([
        ('BACKGROUND',   (0, 0), (-1, -1), accent),
        ('BOX',          (0, 0), (-1, -1), 0.5, border_c),
        ('LINEAFTER',    (0, 0), (4,  1),  0.3, border_c),
        ('TOPPADDING',   (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING',(0, 0), (-1, -1), 5),
        ('LEFTPADDING',  (0, 0), (-1, -1), 7),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('VALIGN',       (0, 0), (-1, -1), 'TOP'),
    ]))
    elements.append(meta_t)
    elements.append(Spacer(1, 0.45 * cm))

    # ────────────────────────────────────────────────────────────────────────
    # 3.  TO / CUSTOMER  +  SUBJECT
    # ────────────────────────────────────────────────────────────────────────
    to_label_cell = Paragraph("<b>To :</b>",
        ps("tol", fontName="Helvetica-Bold", fontSize=9))
    to_body = []
    if q.customer_company:
        to_body.append(Paragraph(f"<b>{q.customer_company}</b>",
            ps("tc", fontName="Helvetica-Bold", fontSize=10, textColor=dark, leading=14)))
    for line in (q.customer_address or "").split("\n"):
        if line.strip():
            to_body.append(Paragraph(line.strip(),
                ps("ta", fontSize=8.5, textColor=gray_c, leading=12)))
    if q.customer_name:
        to_body.append(Paragraph(f"Attn : {q.customer_name}", ps("tn", fontSize=9)))
    if q.customer_email:
        to_body.append(Paragraph(f"Email : {q.customer_email}",
            ps("te", fontSize=8.5, textColor=gray_c)))
    if q.customer_phone:
        to_body.append(Paragraph(f"Phone : {q.customer_phone}",
            ps("tp", fontSize=8.5, textColor=gray_c)))

    to_t = Table([[to_label_cell, to_body]],
        colWidths=[1.4 * cm, usable - 1.4 * cm])
    to_t.setStyle(TableStyle([
        ('VALIGN',       (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING',  (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('TOPPADDING',   (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING',(0, 0), (-1, -1), 0),
    ]))
    elements.append(to_t)
    elements.append(Spacer(1, 0.2 * cm))

    if q.project_name:
        elements.append(Paragraph(
            f"<b>Subject :</b>  {q.project_name}",
            ps("subj", fontSize=9, textColor=dark)))
        elements.append(Spacer(1, 0.2 * cm))

    elements.append(HRFlowable(width=usable, thickness=0.5, color=border_c, spaceAfter=0.3 * cm))

    # ────────────────────────────────────────────────────────────────────────
    # 4.  ITEMS TABLE
    # Col widths must sum exactly to usable (17.0 cm):
    #  S/N | Description      | UOM  | Qty  | Unit Price | Disc% | Amount
    #  0.7 | 6.1              | 1.2  | 1.0  | 3.2        | 1.0   | 3.8
    #  total = 0.7+6.1+1.2+1.0+3.2+1.0+3.8 = 17.0 ✓
    # ────────────────────────────────────────────────────────────────────────
    col_w = [0.7*cm, 6.1*cm, 1.2*cm, 1.0*cm, 3.2*cm, 1.0*cm, 3.8*cm]

    th_s  = ps("th",   fontName="Helvetica-Bold", fontSize=8,   textColor=white, leading=11, alignment=TA_CENTER)
    td_s  = ps("td",   fontSize=8.5, leading=12)
    td_sm = ps("tdsm", fontSize=7.5, textColor=gray_c, leading=10)
    td_c  = ps("tdc",  fontSize=8.5, leading=12, alignment=TA_CENTER)
    td_r  = ps("tdr",  fontSize=8.5, leading=12, alignment=TA_RIGHT)
    td_rb = ps("tdrb", fontName="Helvetica-Bold", fontSize=8.5, leading=12, alignment=TA_RIGHT)

    headers_row = [
        Paragraph("S/N",                 th_s),
        Paragraph("Description / Inventory", th_s),
        Paragraph("UOM",                 th_s),
        Paragraph("Qty",                 th_s),
        Paragraph(f"Unit Price ({cur_str})", th_s),
        Paragraph("Disc%",               th_s),
        Paragraph(f"Amount ({cur_str})", th_s),
    ]

    items          = q.items or []
    rows           = [headers_row]
    row_fills      = []
    subtotal       = 0.0
    total_disc_amt = 0.0
    gross_total    = 0.0

    def fmt_num(v, is_idr):
        if is_idr: return f"{v:,.0f}".replace(",", ".")
        return f"{v:,.2f}"

    is_idr = cur_str == "IDR"

    for i, item in enumerate(items):
        gross, disc_amt, net = calc_item(item)
        subtotal       += net
        total_disc_amt += disc_amt
        gross_total    += gross

        desc_cells = []
        if item.get("description"):
            desc_cells.append(Paragraph(f"<b>{item['description']}</b>", td_s))
        brand_str = "  |  ".join(filter(None, [item.get("brand"), item.get("model")]))
        if brand_str:
            desc_cells.append(Paragraph(brand_str, td_sm))
        if item.get("remarks"):
            desc_cells.append(Paragraph(f"<i>{item['remarks']}</i>", td_sm))

        qty  = float(item.get("qty")        or 0)
        disc = float(item.get("discount")   or 0)
        unit = item.get("unit") or "Unit"
        up   = float(item.get("unit_price") or 0)

        disc_str = f"{disc:.0f}%" if disc > 0 else "-"
        if disc > 0:
            disc_para = Paragraph(f"<b>{disc_str}</b>",
                ps(f"dc{i}", fontSize=8, fontName="Helvetica-Bold",
                   textColor=disc_fc, leading=11, alignment=TA_CENTER))
        else:
            disc_para = Paragraph(disc_str, td_c)

        rows.append([
            Paragraph(str(i + 1),          td_c),
            desc_cells,
            Paragraph(unit,                td_c),
            Paragraph(fmt_num(qty, False).rstrip("0").rstrip(".") if "." in fmt_num(qty,False) else fmt_num(qty,False), td_c),
            Paragraph(fmt_num(up, is_idr), td_r),
            disc_para,
            Paragraph(fmt_num(net, is_idr), td_r),
        ])
        if i % 2 == 1:
            row_fills.append(('BACKGROUND', (0, i + 1), (-1, i + 1), lt_gray))

    # ── Totals rows (span last 2 columns — label col 5, value col 6) ────────
    n_items = len(items)

    def total_row(label_txt, value_txt, bold=False, grand=False, is_disc=False):
        lbl_style = ps(f"tl{label_txt}",
            fontName="Helvetica-Bold",
            fontSize=9 if grand else 8.5,
            textColor=white if grand else (disc_fc if is_disc else gray_c),
            alignment=TA_RIGHT)
        val_style = ps(f"tv{label_txt}",
            fontName="Helvetica-Bold",
            fontSize=9 if grand else 8.5,
            textColor=white if grand else (disc_fc if is_disc else dark),
            alignment=TA_RIGHT)
        return ["", "", "", "", "", Paragraph(label_txt, lbl_style), Paragraph(value_txt, val_style)]

    rows.append(total_row("Sub Total :", fmt_num(subtotal, is_idr)))
    n_subtotal = len(rows) - 1

    # ── Discount row — shown ONLY when there is actual discount ─────────────
    has_discount = total_disc_amt > 0.005
    n_disc = None
    if has_discount:
        disc_disp = f"({fmt_num(total_disc_amt, is_idr)})"
        rows.append(total_row("Discount :", disc_disp, is_disc=True))
        n_disc = len(rows) - 1

    # ── VAT row ──────────────────────────────────────────────────────────────
    vat_pct     = float(q.vat_pct or 11) if (q.vat_include or False) else 0
    vat_include = bool(q.vat_include)
    vat_amt     = subtotal * vat_pct / 100 if vat_include else 0
    grand_total = subtotal + vat_amt

    n_vat = None
    if vat_include:
        rows.append(total_row(f"VAT {vat_pct:.0f}% :", fmt_num(vat_amt, is_idr)))
        n_vat = len(rows) - 1

    rows.append(total_row("TOTAL :", fmt_num(grand_total, is_idr), grand=True))
    n_grand = len(rows) - 1

    # ── Table style ───────────────────────────────────────────────────────────
    tbl_style = [
        # Header row
        ('BACKGROUND',    (0, 0),         (-1, 0),          primary),
        ('TEXTCOLOR',     (0, 0),         (-1, 0),          white),
        ('ALIGN',         (0, 0),         (-1, 0),          'CENTER'),
        # Body
        ('VALIGN',        (0, 0),         (-1, -1),         'TOP'),
        ('TOPPADDING',    (0, 0),         (-1, -1),         5),
        ('BOTTOMPADDING', (0, 0),         (-1, -1),         5),
        ('LEFTPADDING',   (0, 0),         (-1, -1),         5),
        ('RIGHTPADDING',  (0, 0),         (-1, -1),         5),
        # Column alignment (body rows)
        ('ALIGN', (0, 1), (0, n_items), 'CENTER'),   # S/N
        ('ALIGN', (2, 1), (2, n_items), 'CENTER'),   # UOM
        ('ALIGN', (3, 1), (3, n_items), 'CENTER'),   # Qty
        ('ALIGN', (4, 1), (4, n_items), 'RIGHT'),    # Unit Price
        ('ALIGN', (5, 1), (5, n_items), 'CENTER'),   # Disc%
        ('ALIGN', (6, 1), (6, n_items), 'RIGHT'),    # Amount
        # Item rows border
        ('BOX',       (0, 0), (-1, n_items), 0.5, border_c),
        ('INNERGRID', (0, 0), (-1, n_items), 0.25, border_c),
        ('LINEBELOW', (0, n_items), (-1, n_items), 1.5, primary),
        # Subtotal separator
        ('LINEBELOW', (5, n_subtotal), (6, n_subtotal), 0.5, border_c),
        # Grand total
        ('BACKGROUND', (5, n_grand), (6, n_grand), primary),
        ('LINEABOVE',  (5, n_grand), (6, n_grand), 1.5, primary),
        ('LINEBELOW',  (5, n_grand), (6, n_grand), 1.5, primary),
        ('TOPPADDING',    (5, n_grand), (6, n_grand), 7),
        ('BOTTOMPADDING', (5, n_grand), (6, n_grand), 7),
    ] + row_fills

    # Discount row styling
    if n_disc is not None:
        tbl_style += [
            ('BACKGROUND', (5, n_disc), (6, n_disc), disc_bg),
            ('LINEABOVE',  (5, n_disc), (6, n_disc), 0.3, border_c),
            ('LINEBELOW',  (5, n_disc), (6, n_disc), 0.3, border_c),
        ]
    # VAT row styling
    if n_vat is not None:
        tbl_style.append(('LINEBELOW', (5, n_vat), (6, n_vat), 0.3, border_c))

    items_t = Table(rows, colWidths=col_w, repeatRows=1)
    items_t.setStyle(TableStyle(tbl_style))
    elements.append(items_t)
    elements.append(Spacer(1, 0.35 * cm))

    # ────────────────────────────────────────────────────────────────────────
    # 5.  NOTE
    # ────────────────────────────────────────────────────────────────────────
    if q.notes:
        elements.append(Paragraph(f"<i>{q.notes}</i>",
            ps("note", fontSize=8, textColor=gray_c, leading=11)))
        elements.append(Spacer(1, 0.3 * cm))

    elements.append(HRFlowable(width=usable, thickness=0.5, color=border_c, spaceAfter=0.35 * cm))

    # ────────────────────────────────────────────────────────────────────────
    # 6.  GENERAL TERMS & CONDITIONS
    # ────────────────────────────────────────────────────────────────────────
    elements.append(Paragraph("<b><u>General Terms &amp; Conditions :</u></b>",
        ps("tch", fontName="Helvetica-Bold", fontSize=9, textColor=dark)))
    elements.append(Spacer(1, 0.2 * cm))

    vat_note       = f"Include VAT {vat_pct:.0f}%" if vat_include else "Exclude VAT"
    currency_label = f"{cur_str} — {vat_note}"

    def tc_row(lbl, val):
        return [
            Paragraph(f"<b>{lbl}</b>",
                ps(f"tcl{lbl}", fontName="Helvetica-Bold", fontSize=9, textColor=dark)),
            Paragraph(":", ps(f"tcc{lbl}", fontSize=9, alignment=TA_CENTER)),
            Paragraph(str(val), ps(f"tcv{lbl}", fontSize=9)),
        ]

    tc_rows = [
        tc_row("Currency",       currency_label),
        tc_row("Shipment Terms", q.shipment_terms or "-"),
        tc_row("Validity",       "20 DAYS"),
        tc_row("Delivery",       q.delivery       or "-"),
        tc_row("Payment Terms",  q.payment_terms  or "-"),
    ]
    if q.terms and q.terms.strip():
        tc_rows.append(tc_row("Additional Terms", q.terms.replace("\n", "<br/>")))

    tc_t = Table(tc_rows, colWidths=[3.8 * cm, 0.4 * cm, usable - 4.2 * cm])
    tc_t.setStyle(TableStyle([
        ('VALIGN',        (0, 0), (-1, -1), 'TOP'),
        ('TOPPADDING',    (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('LEFTPADDING',   (0, 0), (-1, -1), 0),
        ('RIGHTPADDING',  (0, 0), (-1, -1), 0),
    ]))
    elements.append(tc_t)
    elements.append(Spacer(1, 0.6 * cm))

    # ────────────────────────────────────────────────────────────────────────
    # 7.  REGARDS
    # ────────────────────────────────────────────────────────────────────────
    elements.append(Paragraph("Regards,", ps("reg", fontSize=9, textColor=gray_c)))
    elements.append(Spacer(1, 1.0 * cm))
    if q.sales_person:
        elements.append(Paragraph(f"<b>{q.sales_person}</b>",
            ps("sp", fontName="Helvetica-Bold", fontSize=9.5, textColor=dark)))
    elements.append(Paragraph("PT. Flotech Controls Indonesia",
        ps("spc", fontSize=8.5, textColor=gray_c)))

    # ────────────────────────────────────────────────────────────────────────
    # 8.  FOOTER  (drawn on canvas for every page)
    # ────────────────────────────────────────────────────────────────────────
    def footer_cb(cv, doc_obj):
        cv.saveState()
        pw, ph = A4
        y_line = 3.3 * cm
        # Horizontal rule — aligned with text margins
        cv.setStrokeColor(primary)
        cv.setLineWidth(1)
        cv.line(L, y_line, pw - R, y_line)
        # Company name
        cv.setFont("Helvetica-Bold", 9)
        cv.setFillColor(primary)
        cv.drawCentredString(pw / 2, y_line - 0.45 * cm, FLOTECH_INFO["name"])
        # Address / contact lines
        cv.setFont("Helvetica", 7.5)
        cv.setFillColor(colors.HexColor("#6B7280"))
        cv.drawCentredString(pw / 2, y_line - 0.85 * cm,
            f"{FLOTECH_INFO['address']}  |  {FLOTECH_INFO['city']}")
        cv.drawCentredString(pw / 2, y_line - 1.2 * cm, FLOTECH_INFO["telp"])
        cv.drawCentredString(pw / 2, y_line - 1.55 * cm, FLOTECH_INFO["email"])
        # Page number — right-aligned
        cv.setFont("Helvetica", 7)
        cv.setFillColor(colors.HexColor("#9CA3AF"))
        cv.drawRightString(pw - R, 0.75 * cm, f"Page {doc_obj.page}")
        cv.restoreState()

    doc.build(elements, onFirstPage=footer_cb, onLaterPages=footer_cb)
    buffer.seek(0)
    return buffer

# ── PDF endpoints ──────────────────────────────────────────────────────────────
@quotation_bp.route('/pdf/<int:qid>', methods=['GET'])
@jwt_required()
def quotation_pdf(qid):
    q = Quotation.query.get(qid)
    if not q: return jsonify({"error": "Not found"}), 404
    buf = build_quotation_pdf(q)
    return send_file(buf, as_attachment=True,
        download_name=f"Quotation_{q.quotation_number}.pdf",
        mimetype="application/pdf")

@quotation_bp.route('/pdf/preview/<int:qid>', methods=['GET'])
@jwt_required()
def quotation_pdf_preview(qid):
    q = Quotation.query.get(qid)
    if not q: return jsonify({"error": "Not found"}), 404
    buf = build_quotation_pdf(q)
    return Response(buf, mimetype="application/pdf",
        headers={"Content-Disposition": f"inline; filename=Quotation_{q.quotation_number}.pdf"})# ═══════════════════════════════════════════════════════════════════════════════
# PDF BUILDER — Single Quotation  (A4, margins L/R = 1.8cm → usable = 17.4cm)
# Layout: header | blue divider | meta-band | customer+project | items | terms
# ═══════════════════════════════════════════════════════════════════════════════
def build_quotation_pdf(q):
    buffer = BytesIO()

    # ── Page geometry ────────────────────────────────────────────────────────
    PW, PH  = A4                   # 595.28 × 841.89 pt
    L = R   = 1.8 * cm
    TOP     = 2.0 * cm
    BOT     = 3.8 * cm             # space reserved for footer
    UW      = PW - L - R           # usable width ≈ 17.4 cm

    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        topMargin=TOP, bottomMargin=BOT,
        leftMargin=L, rightMargin=R,
        title=f"Quotation {q.quotation_number}",
        author="PT. Flotech Controls Indonesia",
    )

    # ── Palette ──────────────────────────────────────────────────────────────
    C_PRIMARY  = colors.HexColor("#0B3D91")
    C_NAVY2    = colors.HexColor("#1E5CC6")
    C_ACCENT   = colors.HexColor("#EEF3FB")
    C_DARK     = colors.HexColor("#111827")
    C_TEXT     = colors.HexColor("#374151")
    C_GRAY     = colors.HexColor("#6B7280")
    C_LTGRAY   = colors.HexColor("#F3F4F6")
    C_BORDER   = colors.HexColor("#D1D5DB")
    C_WHITE    = colors.white
    C_DISC_BG  = colors.HexColor("#FFF7ED")
    C_DISC_FC  = colors.HexColor("#B45309")
    C_GRANDROW = colors.HexColor("#0B3D91")

    # ── Style factory ────────────────────────────────────────────────────────
    def S(name, **kw):
        d = dict(fontName="Helvetica", fontSize=9, leading=12, textColor=C_TEXT)
        d.update(kw)
        return ParagraphStyle(name, **d)

    cur    = q.currency or "IDR"
    is_idr = (cur == "IDR")

    def fmt(v):
        n = float(v or 0)
        if is_idr:
            return f"{n:,.0f}".replace(",",".")
        return f"{n:,.2f}"

    elements = []

    # ════════════════════════════════════════════════════════════════════════
    # SECTION 1 — HEADER  (logo left, company+title right)
    # ════════════════════════════════════════════════════════════════════════
    logo_path = _find_logo()
    if logo_path:
        try:
            pil    = PILImage.open(logo_path)
            lw     = 5.2 * cm
            lh     = lw * pil.size[1] / pil.size[0]
            logo_c = RLImage(logo_path, width=lw, height=lh)
        except Exception:
            logo_c = Paragraph("<b>FLOTECH</b>",
                S("lf", fontName="Helvetica-Bold", fontSize=20, textColor=C_PRIMARY))
    else:
        logo_c = Paragraph(
            "<b>FLOTECH</b><br/><font size='7'>PROCESS CONTROL &amp; INSTRUMENTATION</font>",
            S("lf2", fontName="Helvetica-Bold", fontSize=18, textColor=C_PRIMARY, leading=22))

    right_c = [
        Paragraph("PT. FLOTECH CONTROLS INDONESIA",
            S("cn", fontName="Helvetica-Bold", fontSize=9.5, textColor=C_PRIMARY,
              leading=13, alignment=TA_RIGHT)),
        Spacer(1, 2),
        Paragraph("SALES QUOTATION",
            S("qt", fontName="Helvetica-Bold", fontSize=20, textColor=C_PRIMARY,
              leading=24, alignment=TA_RIGHT)),
    ]

    hdr_t = Table([[logo_c, right_c]], colWidths=[6.0*cm, UW - 6.0*cm])
    hdr_t.setStyle(TableStyle([
        ("VALIGN",        (0,0),(-1,-1),"MIDDLE"),
        ("LEFTPADDING",   (0,0),(-1,-1), 0),
        ("RIGHTPADDING",  (0,0),(-1,-1), 0),
        ("TOPPADDING",    (0,0),(-1,-1), 0),
        ("BOTTOMPADDING", (0,0),(-1,-1), 0),
    ]))
    elements.append(hdr_t)
    elements.append(HRFlowable(width=UW, thickness=1, color=C_PRIMARY, spaceBefore=10, spaceAfter=20))

    # ════════════════════════════════════════════════════════════════════════
    # SECTION 2 — META BAND  (6 equal columns, full width)
    # ════════════════════════════════════════════════════════════════════════
    q_date  = q.created_at.strftime("%d %B %Y") if q.created_at else datetime.utcnow().strftime("%d %B %Y")
    rev_str = f"Rev. {q.revision}" if (q.revision or 0) > 0 else "Original"
    cw6     = UW / 6

    def mcell(lbl, val):
        return [
            Paragraph(lbl, S(f"ml{lbl}", fontSize=7, fontName="Helvetica-Bold",
                             textColor=C_GRAY, leading=9, alignment=TA_LEFT)),
            Paragraph(str(val), S(f"mv{lbl}", fontSize=9, fontName="Helvetica-Bold",
                                  textColor=C_DARK, leading=12, alignment=TA_LEFT)),
        ]

    meta_cols = [
        mcell("QUOTATION NO.",  q.quotation_number or "-"),
        mcell("DATE",           q_date),
        mcell("SALES PERSON",   q.sales_person or "-"),
        mcell("REVISION",       rev_str),
        mcell("REF. NO.",       q.ref_no or "-"),
        mcell("CURRENCY",       cur),
    ]

    meta_t = Table(
        [[c[0] for c in meta_cols], [c[1] for c in meta_cols]],
        colWidths=[cw6]*6,
    )
    meta_t.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(-1,-1), C_ACCENT),
        ("BOX",           (0,0),(-1,-1), 0.5, C_BORDER),
        ("LINEAFTER",     (0,0),(4, 1),  0.3, C_BORDER),
        ("TOPPADDING",    (0,0),(-1,-1), 5),
        ("BOTTOMPADDING", (0,0),(-1,-1), 5),
        ("LEFTPADDING",   (0,0),(-1,-1), 7),
        ("RIGHTPADDING",  (0,0),(-1,-1), 4),
        ("VALIGN",        (0,0),(-1,-1), "TOP"),
    ]))
    elements.append(meta_t)
    elements.append(Spacer(1, 0.4*cm))

    # ════════════════════════════════════════════════════════════════════════
    # SECTION 3 — CUSTOMER  (left 56%)  +  PROJECT  (right 44%)
    # ════════════════════════════════════════════════════════════════════════
    GAP    = 0.3 * cm
    cust_w = UW * 0.56 - GAP / 2
    proj_w = UW * 0.44 - GAP / 2

    def cust_block():
        rows = [[Paragraph("CUSTOMER / RECIPIENT",
            S("ch", fontName="Helvetica-Bold", fontSize=7.5,
              textColor=C_WHITE, leading=10))]]
        if q.customer_company:
            rows.append([Paragraph(q.customer_company,
                S("cc", fontName="Helvetica-Bold", fontSize=10.5,
                  textColor=C_DARK, leading=14))])
        for ln in (q.customer_address or "").split("\n"):
            if ln.strip():
                rows.append([Paragraph(ln.strip(),
                    S("ca", fontSize=8.5, textColor=C_GRAY, leading=12))])
        if q.customer_name:
            rows.append([Paragraph(f"Attn : {q.customer_name}",
                S("cn2", fontSize=9, textColor=C_TEXT))])
        if q.customer_email:
            rows.append([Paragraph(f"Email : {q.customer_email}",
                S("ce", fontSize=8.5, textColor=C_GRAY))])
        if q.customer_phone:
            rows.append([Paragraph(f"Phone : {q.customer_phone}",
                S("cp", fontSize=8.5, textColor=C_GRAY))])
        t = Table(rows, colWidths=[cust_w])
        t.setStyle(TableStyle([
            ("BACKGROUND",    (0,0),(0,0),  C_PRIMARY),
            ("BACKGROUND",    (0,1),(-1,-1),C_ACCENT),
            ("BOX",           (0,0),(-1,-1),0.5, C_BORDER),
            ("TOPPADDING",    (0,0),(-1,-1),6),
            ("BOTTOMPADDING", (0,0),(-1,-1),6),
            ("LEFTPADDING",   (0,0),(-1,-1),8),
            ("RIGHTPADDING",  (0,0),(-1,-1),8),
            ("VALIGN",        (0,0),(-1,-1),"MIDDLE"),
        ]))
        return t

    def proj_block():
        rows = [[Paragraph("PROJECT DETAILS",
            S("ph", fontName="Helvetica-Bold", fontSize=7.5,
              textColor=C_WHITE, leading=10))]]
        rows.append([Paragraph(q.project_name or "-",
            S("pn", fontName="Helvetica-Bold", fontSize=10.5,
              textColor=C_DARK, leading=14))])
        if q.category:
            rows.append([Paragraph(f"Category : {q.category}",
                S("pc", fontSize=8.5, textColor=C_GRAY))])
        if q.valid_until:
            rows.append([Paragraph(
                f"Valid Until : {q.valid_until.strftime('%d %B %Y')}",
                S("pv", fontSize=8.5, textColor=C_GRAY))])
        t = Table(rows, colWidths=[proj_w])
        t.setStyle(TableStyle([
            ("BACKGROUND",    (0,0),(0,0),  C_NAVY2),
            ("BACKGROUND",    (0,1),(-1,-1),C_ACCENT),
            ("BOX",           (0,0),(-1,-1),0.5, C_BORDER),
            ("TOPPADDING",    (0,0),(-1,-1),6),
            ("BOTTOMPADDING", (0,0),(-1,-1),6),
            ("LEFTPADDING",   (0,0),(-1,-1),8),
            ("RIGHTPADDING",  (0,0),(-1,-1),8),
            ("VALIGN",        (0,0),(-1,-1),"MIDDLE"),
        ]))
        return t

    two_col = Table([[cust_block(), proj_block()]],
        colWidths=[cust_w + GAP/2, proj_w + GAP/2])
    two_col.setStyle(TableStyle([
        ("VALIGN",        (0,0),(-1,-1),"TOP"),
        ("LEFTPADDING",   (0,0),(-1,-1),0),
        ("RIGHTPADDING",  (0,0),(-1,-1),0),
        ("TOPPADDING",    (0,0),(-1,-1),0),
        ("BOTTOMPADDING", (0,0),(-1,-1),0),
        ("RIGHTPADDING",  (0,0),(0,0),  GAP/2),
        ("LEFTPADDING",   (1,0),(1,0),  GAP/2),
    ]))
    elements.append(two_col)
    elements.append(Spacer(1, 0.4*cm))

    # ════════════════════════════════════════════════════════════════════════
    # SECTION 4 — ITEMS TABLE
    # Column widths must sum to UW (≈ 17.4 cm = PW - 2×1.8cm)
    #   S/N  | Description        | UOM  | Qty  | Unit Price  | Disc% | Amount
    #   0.75 | 6.55               | 1.2  | 1.0  | 3.3         | 1.0   | 3.6
    # ════════════════════════════════════════════════════════════════════════
    CW = [0.75*cm, 6.55*cm, 1.2*cm, 1.0*cm, 3.3*cm, 1.0*cm, 3.6*cm]
    # Adjust last col so sum == UW exactly
    CW[-1] = UW - sum(CW[:-1])

    TH  = S("th",  fontName="Helvetica-Bold", fontSize=8,   textColor=C_WHITE, leading=11, alignment=TA_CENTER)
    TD  = S("td",  fontSize=8.5, leading=12)
    TDS = S("tds", fontSize=7.5, textColor=C_GRAY, leading=10)
    TDC = S("tdc", fontSize=8.5, leading=12, alignment=TA_CENTER)
    TDR = S("tdr", fontSize=8.5, leading=12, alignment=TA_RIGHT)

    rows      = [[
        Paragraph("S/N",                     TH),
        Paragraph("Description / Inventory", TH),
        Paragraph("UOM",                     TH),
        Paragraph("Qty",                     TH),
        Paragraph(f"Unit Price\n({cur})",    TH),
        Paragraph("Disc\n%",                 TH),
        Paragraph(f"Amount\n({cur})",        TH),
    ]]
    row_fills  = []
    subtotal   = 0.0
    total_disc = 0.0
    items      = q.items or []

    for i, item in enumerate(items):
        gross, d_amt, net = calc_item(item)
        subtotal   += net
        total_disc += d_amt

        desc_flow = []
        if item.get("description"):
            desc_flow.append(Paragraph(f"<b>{item['description']}</b>", TD))
        bm = "  |  ".join(filter(None, [item.get("brand"), item.get("model")]))
        if bm:
            desc_flow.append(Paragraph(bm, TDS))
        if item.get("remarks"):
            desc_flow.append(Paragraph(f"<i>{item['remarks']}</i>", TDS))

        qty   = float(item.get("qty")        or 0)
        disc  = float(item.get("discount")   or 0)
        up    = float(item.get("unit_price") or 0)
        unit  = item.get("unit") or "Unit"
        qty_s = str(int(qty)) if qty == int(qty) else f"{qty:g}"

        disc_p = (
            Paragraph(f"<b>{disc:.0f}%</b>",
                S(f"dc{i}", fontName="Helvetica-Bold", fontSize=8,
                  textColor=C_DISC_FC, leading=11, alignment=TA_CENTER))
            if disc > 0 else Paragraph("—", TDC)
        )

        rows.append([
            Paragraph(str(i+1), TDC),
            desc_flow,
            Paragraph(unit,     TDC),
            Paragraph(qty_s,    TDC),
            Paragraph(fmt(up),  TDR),
            disc_p,
            Paragraph(fmt(net), TDR),
        ])
        if i % 2 == 1:
            row_fills.append(("BACKGROUND", (0,i+1),(-1,i+1), C_LTGRAY))

    n_items = len(items)

    # ── Total subtable (same width as items table, pinned right) ─────────────
    vat_pct     = float(q.vat_pct or 11) if (q.vat_include or False) else 0
    vat_include = bool(q.vat_include)
    vat_amt     = subtotal * vat_pct / 100 if vat_include else 0
    grand_total = subtotal + vat_amt
    has_disc    = total_disc > 0.005

    # Total label col = first 5 item cols width; value col = last 2
    lbl_w = sum(CW[:5])   # ≈ 12.8 cm
    val_w = sum(CW[5:])   # ≈  4.6 cm

    SL = S("sl", fontName="Helvetica-Bold", fontSize=9,  textColor=C_GRAY,    alignment=TA_RIGHT)
    SV = S("sv", fontName="Helvetica-Bold", fontSize=9,  textColor=C_DARK,    alignment=TA_RIGHT)
    GL = S("gl", fontName="Helvetica-Bold", fontSize=10, textColor=C_WHITE,   alignment=TA_RIGHT)
    GV = S("gv", fontName="Helvetica-Bold", fontSize=10, textColor=C_WHITE,   alignment=TA_RIGHT)
    DL = S("dl", fontName="Helvetica-Bold", fontSize=9,  textColor=C_DISC_FC, alignment=TA_RIGHT)
    DV = S("dv", fontName="Helvetica-Bold", fontSize=9,  textColor=C_DISC_FC, alignment=TA_RIGHT)

    tot_rows   = []
    tot_styles = [
        ("LEFTPADDING",   (0,0),(-1,-1), 8),
        ("RIGHTPADDING",  (0,0),(-1,-1), 8),
        ("TOPPADDING",    (0,0),(-1,-1), 5),
        ("BOTTOMPADDING", (0,0),(-1,-1), 5),
        ("VALIGN",        (0,0),(-1,-1), "MIDDLE"),
        ("LINEBELOW",     (0,0),(-1, 0), 0.5, C_BORDER),
    ]

    tot_rows.append([Paragraph("Sub Total (IDR) :", SL), Paragraph(fmt(subtotal), SV)])

    if has_disc:
        r = len(tot_rows)
        tot_rows.append([Paragraph("Discount :", DL),
                         Paragraph(f"({fmt(total_disc)})", DV)])
        tot_styles += [
            ("BACKGROUND", (0,r),(-1,r), C_DISC_BG),
            ("LINEABOVE",  (0,r),(-1,r), 0.5, C_BORDER),
            ("LINEBELOW",  (0,r),(-1,r), 0.5, C_BORDER),
        ]

    if vat_include:
        r = len(tot_rows)
        tot_rows.append([Paragraph(f"VAT {vat_pct:.0f}% :", SL),
                         Paragraph(fmt(vat_amt), SV)])
        tot_styles.append(("LINEBELOW", (0,r),(-1,r), 0.3, C_BORDER))

    r_grand = len(tot_rows)
    tot_rows.append([Paragraph("GRAND TOTAL (IDR) :", GL), Paragraph(fmt(grand_total), GV)])
    tot_styles += [
        ("BACKGROUND",    (0,r_grand),(-1,r_grand), C_PRIMARY),
        ("TOPPADDING",    (0,r_grand),(-1,r_grand), 7),
        ("BOTTOMPADDING", (0,r_grand),(-1,r_grand), 7),
    ]

    tot_t = Table(tot_rows, colWidths=[lbl_w, val_w])
    tot_t.setStyle(TableStyle(tot_styles))

    tbl_style = [
        ("BACKGROUND",    (0,0),(-1,0),         C_PRIMARY),
        ("TEXTCOLOR",     (0,0),(-1,0),          C_WHITE),
        ("ALIGN",         (0,0),(-1,0),          "CENTER"),
        ("VALIGN",        (0,0),(-1,-1),         "TOP"),
        ("TOPPADDING",    (0,0),(-1,-1),         5),
        ("BOTTOMPADDING", (0,0),(-1,-1),         5),
        ("LEFTPADDING",   (0,0),(-1,-1),         5),
        ("RIGHTPADDING",  (0,0),(-1,-1),         5),
        ("ALIGN",  (0,1),(0,n_items), "CENTER"),
        ("ALIGN",  (2,1),(2,n_items), "CENTER"),
        ("ALIGN",  (3,1),(3,n_items), "CENTER"),
        ("ALIGN",  (4,1),(4,n_items), "RIGHT"),
        ("ALIGN",  (5,1),(5,n_items), "CENTER"),
        ("ALIGN",  (6,1),(6,n_items), "RIGHT"),
        ("BOX",      (0,0),(-1,n_items), 0.5, C_BORDER),
        ("INNERGRID",(0,0),(-1,n_items), 0.25, C_BORDER),
        ("LINEBELOW",(0,n_items),(-1,n_items), 2.0, C_PRIMARY),
    ] + row_fills

    items_t = Table(rows, colWidths=CW, repeatRows=1)
    items_t.setStyle(TableStyle(tbl_style))

    elements.append(items_t)
    elements.append(tot_t)
    elements.append(Spacer(1, 0.35*cm))

    # ════════════════════════════════════════════════════════════════════════
    # SECTION 5 — NOTE
    # ════════════════════════════════════════════════════════════════════════
    if q.notes and q.notes.strip():
        elements.append(Paragraph(
            f"<i>Note : {q.notes}</i>",
            S("note", fontSize=8, textColor=C_GRAY, leading=11)))
        elements.append(Spacer(1, 0.25*cm))

    elements.append(HRFlowable(width=UW, thickness=0.5, color=C_BORDER, spaceAfter=0.3*cm))

    # ════════════════════════════════════════════════════════════════════════
    # SECTION 6 — TERMS  (left)  +  REGARDS  (right)  — side by side
    # ════════════════════════════════════════════════════════════════════════
    vat_note = f"Include VAT {vat_pct:.0f}%" if vat_include else "Exclude VAT"

    TLH = S("tlh", fontName="Helvetica-Bold", fontSize=8.5, textColor=C_DARK)
    TLV = S("tlv", fontSize=8.5, textColor=C_TEXT, leading=12)
    COL = S("col", fontSize=8.5, textColor=C_GRAY,  alignment=TA_CENTER)

    def tc_row(lbl, val):
        return [
            Paragraph(f"<b>{lbl}</b>", TLH),
            Paragraph(":",             COL),
            Paragraph(str(val),        TLV),
        ]

    tc_data = [
        tc_row("Currency",       f"{cur}  —  {vat_note}"),
        tc_row("Shipment Terms", q.shipment_terms or "-"),
        tc_row("Validity",       "20 Days from quotation date"),
        tc_row("Delivery",       q.delivery      or "-"),
        tc_row("Payment Terms",  q.payment_terms or "-"),
    ]
    if q.terms and q.terms.strip():
        tc_data.append(tc_row("Additional Terms",
                               q.terms.replace("\n", "<br/>")))

    tc_lbl_w  = 3.4 * cm
    tc_colon  = 0.4 * cm
    tc_val_w  = UW * 0.58 - tc_lbl_w - tc_colon
    tc_t = Table(tc_data, colWidths=[tc_lbl_w, tc_colon, tc_val_w])
    tc_t.setStyle(TableStyle([
        ("VALIGN",        (0,0),(-1,-1),"TOP"),
        ("TOPPADDING",    (0,0),(-1,-1),2),
        ("BOTTOMPADDING", (0,0),(-1,-1),2),
        ("LEFTPADDING",   (0,0),(-1,-1),0),
        ("RIGHTPADDING",  (0,0),(-1,-1),0),
    ]))

    sig_items = [
        Paragraph("Regards,", S("rg", fontSize=8.5, textColor=C_GRAY)),
        Spacer(1, 1.0*cm),
    ]
    if q.sales_person:
        sig_items.append(Paragraph(f"<u><b>{q.sales_person}</b></u>",
            S("sp", fontName="Helvetica-Bold", fontSize=9.5, textColor=C_DARK)))
    sig_items.append(Paragraph("PT. Flotech Controls Indonesia",
        S("spc", fontSize=8.5, textColor=C_GRAY)))

    terms_w = tc_lbl_w + tc_colon + tc_val_w
    sig_w   = UW - terms_w
    side_t  = Table([[tc_t, sig_items]], colWidths=[terms_w, sig_w])
    side_t.setStyle(TableStyle([
        ("VALIGN",        (0,0),(-1,-1),"TOP"),
        ("LEFTPADDING",   (0,0),(-1,-1),0),
        ("RIGHTPADDING",  (0,0),(-1,-1),0),
        ("TOPPADDING",    (0,0),(-1,-1),0),
        ("BOTTOMPADDING", (0,0),(-1,-1),0),
        ("LEFTPADDING",   (1,0),(1, 0), 0.5*cm),
    ]))

    elements.append(Paragraph("<b><u>General Terms &amp; Conditions</u></b>",
        S("tch", fontName="Helvetica-Bold", fontSize=9, textColor=C_DARK)))
    elements.append(Spacer(1, 0.2*cm))
    elements.append(side_t)

    # ════════════════════════════════════════════════════════════════════════
    # FOOTER  (canvas callback)
    # ════════════════════════════════════════════════════════════════════════
    def footer_cb(cv, doc_obj):
        cv.saveState()
        y = BOT - 0.2 * cm
        cv.setStrokeColor(C_PRIMARY); cv.setLineWidth(1)
        cv.line(L, y, PW - R, y)
        cv.setFont("Helvetica-Bold", 9); cv.setFillColor(C_PRIMARY)
        cv.drawCentredString(PW/2, y - 0.45*cm, FLOTECH_INFO["name"])
        cv.setFont("Helvetica", 7.5); cv.setFillColor(colors.HexColor("#6B7280"))
        cv.drawCentredString(PW/2, y - 0.82*cm,
            f"{FLOTECH_INFO['address']}  |  {FLOTECH_INFO['city']}")
        cv.drawCentredString(PW/2, y - 1.14*cm, FLOTECH_INFO["telp"])
        cv.drawCentredString(PW/2, y - 1.46*cm, FLOTECH_INFO["email"])
        cv.setFont("Helvetica", 7); cv.setFillColor(colors.HexColor("#9CA3AF"))
        cv.drawRightString(PW - R, 0.6*cm, f"Page {doc_obj.page}")
        cv.restoreState()

    doc.build(elements, onFirstPage=footer_cb, onLaterPages=footer_cb)
    buffer.seek(0)
    return buffer